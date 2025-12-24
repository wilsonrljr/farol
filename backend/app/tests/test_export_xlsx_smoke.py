from __future__ import annotations

from io import BytesIO

from fastapi.testclient import TestClient

from backend.app.main import app


client = TestClient(app)


def test_export_compare_scenarios_enhanced_xlsx_smoke():
    payload = {
        "property_value": 300_000.0,
        "down_payment": 60_000.0,
        "total_savings": 100_000.0,
        "loan_term_years": 2,
        "annual_interest_rate": 9.6,
        "loan_type": "PRICE",
        "rent_value": 1_000.0,
        "investment_returns": [{"start_month": 1, "annual_rate": 0.0}],
        "additional_costs": {
            "itbi_percentage": 2.0,
            "deed_percentage": 1.0,
            "monthly_hoa": 0.0,
            "monthly_property_tax": 0.0,
        },
        "inflation_rate": 0.0,
        "rent_inflation_rate": 0.0,
        "property_appreciation_rate": 0.0,
        "rent_reduces_investment": False,
        "invest_external_surplus": False,
    }

    r = client.post(
        "/api/compare-scenarios-enhanced/export?format=xlsx&shape=long",
        json=payload,
    )
    assert r.status_code == 200, r.text

    # Basic content checks (avoid overfitting exact bytes).
    assert r.headers.get("content-type", "").startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ) or r.headers.get("content-type", "").startswith("application/octet-stream")
    assert len(r.content) > 1000

    # Verify workbook structure.
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(r.content), read_only=True, data_only=True)
    sheets = set(wb.sheetnames)

    # These are the contract sheets produced by the export endpoint.
    assert "metrics" in sheets
    assert "monthly_long" in sheets
    assert "comparative_summary" in sheets
